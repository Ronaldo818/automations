from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PlayTimeout
from openpyxl import load_workbook, Workbook

# ======================================================
# CONFIG
# ======================================================

URL_LOGIN = "https://sistema.avivar.com.br/webpro/webpad/acesso"

USUARIO   = "rogjunio"
SENHA     = "Brasil20@"

# Planilha base (mesma do processo anterior) e LOG específico de rotas
XLSX_BASE       = r"C:\Users\ronaldo.gontijo\Downloads\Clientes.xlsx"            # deve conter colunas: Cliente, Rota
XLSX_LOG_ROTAS  = r"C:\Users\ronaldo.gontijo\Downloads\clientes_rotas_log.xlsx"  # log deste processo

# Módulo e unidade (ajuste conforme necessário)
MODULO  = "22776"
UNIDADE = "1"

# Frame principal
FRAME_NAME = "frameprog"

# Selectors (Central / Manutenção / Rotas)
SEL_PESQ_INPUT    = "#vpad-pesq"
SEL_PESQ_BTN      = "#vpad-btpesq\\.x"
SEL_TABELA        = "#vtabela"

SEL_BTN_CLIENTE   = "input[name='btcliente.x'][value='Cliente']"

SEL_ABA_COMERCIAL = "#vpad-btaba1"           # Dados Comerciais
SEL_ABA_LOGISTICA = "#vpad-btaba4"           # Dados de Logística
CLASS_ABA_ATIVA   = "activeaba"

SEL_ROTA_INPUT    = "#vrot-codigo"

SEL_SALVAR        = "input[name='vpad-btsal.x'][value='Salvar']"
SEL_RADIOS        = "input.radio_sel"

# Tolerâncias (ms)
TIMEOUT_SHORT   = 8000
TIMEOUT_DEFAULT = 20000
TIMEOUT_LONG    = 45000

# Desempenho (esperas rápidas)
FAST_POLL_MS     = 300
FAST_WINDOW_MS   = 3000     # janela rápida (3s) para detecção de retorno
NETIDLE_FAST_MS  = 1500     # networkidle curto após Enter
SLOW_FALLBACK_MS = 8000     # fallback moderado

# Log header
CABECALHO_LOG_ROTAS = [
    "Cliente",
    "Rota",
    "Resultado",
    "Mensagem",
    "DataHoraExecucao"
]

# Screenshots (opcional; deixe "" para desabilitar)
SCREENSHOTS_DIR = r""  # ex: r"C:\temp\screens\rotas"


# ======================================================
# FS / XLSX
# ======================================================

def ensure_parent_dir(path: str):
    Path(path).resolve().parent.mkdir(parents=True, exist_ok=True)

def criar_log_rotas_se_nao_existir(path: str):
    ensure_parent_dir(path)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "LogRotas"
        ws.append(CABECALHO_LOG_ROTAS)
        wb.save(path)
        wb.close()

def salvar_log_rota(path: str, cliente: str, rota: str, resultado: str, mensagem: str):
    try:
        wb = load_workbook(path)
    except FileNotFoundError:
        criar_log_rotas_se_nao_existir(path)
        wb = load_workbook(path)
    except PermissionError as e:
        print(f"[LOG] PermissionError ao abrir o log: {e}")
        return

    try:
        ws = wb.active
        ws.append([
            cliente,
            rota,
            resultado,
            mensagem,
            datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        ])
        wb.save(path)
    except PermissionError as e:
        print(f"[LOG] PermissionError ao salvar o log: {e}")
    finally:
        try: wb.close()
        except: pass

def carregar_base_clientes_e_rotas(path: str):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Base não encontrada: {path}")

    wb = load_workbook(path)
    ws = wb.active

    # Lê cabeçalho
    cabec = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if "Cliente" not in cabec or "Rota" not in cabec:
        wb.close()
        raise ValueError("A planilha base precisa ter as colunas 'Cliente' e 'Rota'.")

    cliente_idx = cabec.index("Cliente")
    rota_idx    = cabec.index("Rota")

    dados = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cliente = row[cliente_idx] if len(row) > cliente_idx else None
        rota    = row[rota_idx] if len(row) > rota_idx else None
        if cliente:
            dados.append({
                "Cliente": str(cliente).strip(),
                "Rota": str(rota).strip() if rota is not None else ""
            })
    wb.close()
    return dados


# ======================================================
# NORMALIZAÇÃO
# ======================================================

def normalizar_rota(valor: str) -> str:
    """
    Mantém apenas dígitos, máximo 6 (maxlength=6).
    """
    if not valor:
        return ""
    s = re.sub(r"\D", "", str(valor))
    return s[:6]


# ======================================================
# SCREENSHOTS
# ======================================================

def ensure_screens_dir():
    if SCREENSHOTS_DIR:
        Path(SCREENSHOTS_DIR).mkdir(parents=True, exist_ok=True)

def take_screenshot(page, nome_prefixo: str):
    if not SCREENSHOTS_DIR:
        return
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    path = os.path.join(SCREENSHOTS_DIR, f"{nome_prefixo}_{ts}.png")
    try:
        page.screenshot(path=path, full_page=True)
        print(f"[Screenshot] {path}")
    except Exception:
        pass


# ======================================================
# PLAYWRIGHT HELPERS
# ======================================================

def entrar_frame(page):
    page.wait_for_selector(f"iframe[name='{FRAME_NAME}']", timeout=TIMEOUT_DEFAULT)
    return page.frame(name=FRAME_NAME)

def abrir_central(page):
    page.click("a:has-text('Manutenções')")
    page.click("a:has-text('Cadastros')")
    page.click("a:has-text('Clientes')")
    page.click("a:has-text('Central Cadastro Cliente')")

    frame = entrar_frame(page)
    frame.wait_for_selector(SEL_PESQ_INPUT, timeout=TIMEOUT_DEFAULT)
    return frame

def pesquisar_cliente(frame, codigo: str):
    frame.fill(SEL_PESQ_INPUT, str(codigo).strip())
    frame.click(SEL_PESQ_BTN)
    frame.wait_for_selector(SEL_TABELA, timeout=TIMEOUT_DEFAULT)

def abrir_manutencao_via_botao_cliente_com_fallback(frame):
    """
    Tenta clicar no botão 'Cliente' diretamente (pois a linha já vem selecionada após a pesquisa).
    Fallback: se falhar, seleciona o primeiro radio da tabela e tenta novamente.
    """
    try:
        frame.click(SEL_BTN_CLIENTE, timeout=TIMEOUT_SHORT)
        frame.wait_for_selector(SEL_SALVAR, timeout=TIMEOUT_DEFAULT)
        return
    except Exception:
        pass

    radios = frame.locator(SEL_RADIOS)
    if radios.count() > 0:
        try:
            radios.first.click()
            frame.click(SEL_BTN_CLIENTE, timeout=TIMEOUT_SHORT)
            frame.wait_for_selector(SEL_SALVAR, timeout=TIMEOUT_DEFAULT)
            return
        except Exception:
            pass

    # Última tentativa
    frame.click(SEL_BTN_CLIENTE, timeout=TIMEOUT_LONG)
    frame.wait_for_selector(SEL_SALVAR, timeout=TIMEOUT_DEFAULT)

def garantir_aba_logistica_ativa(frame):
    """
    Garante que a aba 'Dados de Logística' (#vpad-btaba4) está ativa.
    Alguns ambientes alternam automaticamente para 'Dados Comerciais' (#vpad-btaba1) após blur/Enter.
    """
    try:
        aba_log = frame.locator(SEL_ABA_LOGISTICA)
        if aba_log.count() > 0:
            cls = aba_log.first.get_attribute("class") or ""
            if CLASS_ABA_ATIVA in cls:
                return  # já está ativa
        frame.click(SEL_ABA_LOGISTICA)
        frame.wait_for_selector(SEL_ROTA_INPUT, timeout=TIMEOUT_DEFAULT)
    except Exception:
        frame.wait_for_selector(SEL_ROTA_INPUT, timeout=TIMEOUT_DEFAULT)

def abrir_aba_logistica(frame):
    frame.click(SEL_ABA_LOGISTICA)
    frame.wait_for_selector(SEL_ROTA_INPUT, timeout=TIMEOUT_DEFAULT)

def preencher_rota_com_enter(page, frame, rota: str):
    """
    Preenche a rota, envia ENTER (para disparar fbusca/consulta), aguarda carregar,
    garante permanecer na aba 'Dados de Logística'.
    Também trata o caso em que o ENTER já retorna à lista (nesse caso, pulamos o Salvar).
    """
    garantir_aba_logistica_ativa(frame)

    campo = frame.locator(SEL_ROTA_INPUT)
    campo.fill(rota)

    # ENTER no campo para disparar onblur/consulta
    campo.press("Enter")

    # Aguardar carregamento/validação rápida
    try:
        page.wait_for_load_state("networkidle", timeout=NETIDLE_FAST_MS)
    except Exception:
        pass

    # Se ENTER já levou para a lista, devolve um sinal (retornando None); o chamador vai detectar.
    fr = entrar_frame(page)
    if fr.locator(SEL_TABELA).count() > 0:
        return None

    # Caso contrário, reassegura a aba logística e o campo de rota
    garantir_aba_logistica_ativa(fr)
    fr.wait_for_selector(SEL_ROTA_INPUT, timeout=TIMEOUT_DEFAULT)
    return fr

def salvar_e_voltar_lista(page, frame):
    """
    Salva e aguarda retornar para a lista.
    - Garante aba logística ativa antes do clique
    - scroll + force=True
    - espera competitiva curta para '#vtabela'
    - se a tela cair na aba 1, reativa aba 4 e clica Salvar **uma única vez** como fallback.
    """
    # 0) Se já estamos na lista, nada a fazer
    fr0 = entrar_frame(page)
    if fr0.locator(SEL_TABELA).count() > 0:
        return fr0

    # 1) garante aba logística ativa
    garantir_aba_logistica_ativa(frame)

    # 2) clicar salvar (robusto)
    btn = frame.locator(SEL_SALVAR)
    try:
        btn.scroll_into_view_if_needed()
    except Exception:
        pass
    btn.click(force=True)

    # 3) janela rápida para detectar lista
    total_checks = int(FAST_WINDOW_MS / FAST_POLL_MS)
    second_try_done = False
    for _ in range(total_checks):
        page.wait_for_timeout(FAST_POLL_MS)
        fr = entrar_frame(page)
        if fr.locator(SEL_TABELA).count() > 0:
            return fr
        # se a tela "pular" para aba 1, reative a aba 4 e force mais um clique no salvar (apenas 1x)
        if not second_try_done:
            aba1 = fr.locator(SEL_ABA_COMERCIAL)
            if aba1.count() > 0:
                cls = aba1.first.get_attribute("class") or ""
                if CLASS_ABA_ATIVA in cls:
                    try:
                        garantir_aba_logistica_ativa(fr)
                        b2 = fr.locator(SEL_SALVAR)
                        b2.scroll_into_view_if_needed()
                        b2.click(force=True)
                        second_try_done = True
                    except Exception:
                        pass

    # 4) estabilização curta + checagem final
    try:
        page.wait_for_load_state("networkidle", timeout=NETIDLE_FAST_MS)
    except Exception:
        pass

    fr = entrar_frame(page)
    fr.wait_for_selector(SEL_TABELA, timeout=SLOW_FALLBACK_MS)
    return fr


# ======================================================
# MAIN
# ======================================================

def main():
    ensure_screens_dir()
    criar_log_rotas_se_nao_existir(XLSX_LOG_ROTAS)
    base = carregar_base_clientes_e_rotas(XLSX_BASE)

    with sync_playwright() as p:
        browser  = p.chromium.launch(headless=False)
        context  = browser.new_context()
        page     = context.new_page()

        # Dialogs
        def on_dialog(dialog):
            try:
                print(f"[DIALOG] {dialog.type}: {dialog.message}")
                dialog.accept()
            except Exception:
                pass
        page.on("dialog", on_dialog)

        # LOGIN
        page.goto(URL_LOGIN)
        page.fill("input[name='vusuario']", USUARIO)
        page.fill("input[name='vsenha']", SENHA)
        page.click("button.BtLogin")
        page.wait_for_load_state("networkidle", timeout=TIMEOUT_LONG)

        # Módulo / Unidade
        page.select_option("#vobj-modulo", MODULO)
        page.wait_for_load_state("networkidle", timeout=TIMEOUT_DEFAULT)
        page.select_option("#vobj-unidade", UNIDADE)
        page.wait_for_load_state("networkidle", timeout=TIMEOUT_DEFAULT)

        # Central
        frame = abrir_central(page)

        # Loop base
        for item in base:
            codigo = item["Cliente"]
            rota   = normalizar_rota(item["Rota"])

            print(f"\n=== CADASTRANDO ROTA do CLIENTE {codigo} (Rota={rota or 'VAZIA'}) ===")

            # Se rota vazia -> pular e logar
            if not rota:
                print(f"[PULADO] Cliente {codigo} sem valor de Rota.")
                salvar_log_rota(XLSX_LOG_ROTAS, codigo, rota, "PULADO", "Rota vazia na planilha base")
                continue

            try:
                # 1) Pesquisar cliente
                pesquisar_cliente(frame, codigo)

                # 2) Ir direto no botão Cliente (fallback: selecionar primeiro radio e tentar de novo)
                abrir_manutencao_via_botao_cliente_com_fallback(frame)

                # 3) Aba Dados de Logística -> preencher rota (com ENTER e aguardo)
                abrir_aba_logistica(frame)
                fr_after_enter = preencher_rota_com_enter(page, frame, rota)

                # Se Enter já retornou pra lista, não precisa salvar
                if fr_after_enter is None:
                    frame = entrar_frame(page)  # já deve estar na lista
                else:
                    # 4) Salvar e voltar para a lista (com reforço de aba)
                    frame = salvar_e_voltar_lista(page, fr_after_enter)

                # 5) Log OK
                salvar_log_rota(XLSX_LOG_ROTAS, codigo, rota, "OK", "Rota cadastrada/atualizada com sucesso")

            except Exception as e:
                print(f"ERRO rota cliente {codigo}: {e}")
                take_screenshot(page, f"rota_erro_{codigo}")

                salvar_log_rota(XLSX_LOG_ROTAS, codigo, rota, "ERRO", str(e))

                # Recuperação: se não estiver na lista, tentativa de voltar
                try:
                    fr2 = entrar_frame(page)
                    if fr2.locator(SEL_TABELA).count() == 0:
                        page.go_back()
                        page.wait_for_load_state("networkidle", timeout=NETIDLE_FAST_MS)
                        frame = entrar_frame(page)
                        frame.wait_for_selector(SEL_TABELA, timeout=SLOW_FALLBACK_MS)
                    else:
                        frame = fr2
                except Exception:
                    frame = abrir_central(page)

        print("\nProcesso de cadastro de rotas concluído.")
        browser.close()


if __name__ == "__main__":
    main()