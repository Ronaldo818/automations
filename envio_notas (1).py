from __future__ import annotations

import os
from datetime import datetime
from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook


URL_LOGIN = "https://sistema.avivar.com.br/webpro/webpad/acesso"
XLSX_PATH = r"C:\Users\ronaldo.gontijo\Downloads\envio_notas.xlsx"


CABECALHO = [
    "Codigo",
    "Numero",
    "Cliente",
    "Data Emissao",
    "Resultado",
    "Mensagem Retorno",
    "DataHoraExecucao"
]


def criar_arquivo_se_nao_existir(path):

    if not os.path.exists(path):

        wb = Workbook()
        ws = wb.active
        ws.title = "EnvioNotas"

        ws.append(CABECALHO)

        wb.save(path)
        wb.close()


def carregar_notas_processadas(path):

    notas = set()

    if not os.path.exists(path):
        return notas

    wb = load_workbook(path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):

        if row[0] and row[1]:
            chave = f"{row[0]}_{row[1]}"
            notas.add(chave)

    wb.close()

    return notas


def salvar_xlsx(path, linha_dict):

    wb = load_workbook(path)
    ws = wb.active

    ws.append([
        linha_dict["Codigo"],
        linha_dict["Numero"],
        linha_dict["Cliente"],
        linha_dict["Data Emissao"],
        linha_dict["Resultado"],
        linha_dict["Mensagem Retorno"],
        linha_dict["DataHoraExecucao"]
    ])

    wb.save(path)
    wb.close()


def executar_envio(usuario, senha, data_ini, data_fim, log_callback=None, controle=None):

    enviados = 0
    erros = 0

    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    criar_arquivo_se_nao_existir(XLSX_PATH)
    notas_processadas = carregar_notas_processadas(XLSX_PATH)

    with sync_playwright() as p:

        browser = p.chromium.launch(
            headless=False,
            args=["--start-maximized"],
            slow_mo=20
        )

        context = browser.new_context(viewport={"width":1920,"height":1080})
        page = context.new_page()

        page.goto(URL_LOGIN)

        page.fill("input[name='vusuario']", usuario)
        page.fill("input[name='vsenha']", senha)

        page.click("button.BtLogin")

        page.wait_for_load_state("networkidle")

        with page.expect_navigation():
            page.select_option("#vobj-modulo","22776")

        with page.expect_navigation():
            page.select_option("#vobj-unidade","2")

        page.click("#ui-id-1")
        page.click("#ui-id-28")
        page.click("#ui-id-29")

        frame = page.frame(name="frameprog")

        frame.wait_for_selector("#vpar-dt-ini")

        frame.fill("#vpar-dt-ini",data_ini)
        frame.fill("#vpar-dt-fim",data_fim)

        with frame.expect_navigation():
            frame.select_option("#wnfs-status","1")

        with frame.expect_navigation():
            frame.click("#vpad-btpesq\\.x")

        tabela = frame.locator("#vtabela")

        frame.wait_for_selector("input.radio_sel")

        texto_primeira = tabela.inner_text()

        total_primeira = frame.locator("input.radio_sel").count()

        modo_primeira = True

        if total_primeira >= 20:

            btn_ult = frame.locator("input[name='vpad-btult']")

            if btn_ult.count() > 0:

                btn_ult.click()

                frame.wait_for_selector("input.radio_sel")

                texto_ultima = tabela.inner_text()

                if texto_primeira != texto_ultima:

                    log("Última aba diferente — trabalhando da última")

                    modo_primeira = False

                else:

                    log("Última aba igual à primeira — trabalhando na primeira")

                    modo_primeira = True

        log("Iniciando processamento...")

        while True:

            if controle:

                executando = controle()

                if not executando:
                    log("Processo encerrado pelo usuário")
                    return

            frame.wait_for_selector("input.radio_sel")

            radios = frame.locator("input.radio_sel")

            total = radios.count()

            encontrou = False

            for i in reversed(range(total)):

                if controle:

                    executando = controle()

                    if not executando:
                        log("Processo encerrado pelo usuário")
                        return

                linha = radios.nth(i).locator("xpath=ancestor::tr")

                codigo = linha.locator("td").nth(1).inner_text().strip()
                numero = linha.locator("td").nth(3).inner_text().strip()

                cliente = linha.locator("td").nth(15).inner_text().strip()
                data_emissao = linha.locator("td").nth(7).inner_text().strip()

                img = linha.locator("img[alt]")

                if img.count() == 0:
                    continue

                if img.first.get_attribute("alt") != "Autorizada":
                    continue

                chave = f"{codigo}_{numero}"

                if chave in notas_processadas:
                    continue

                encontrou = True

                radios.nth(i).click()

                log(f"Enviando nota {numero}")

                with frame.expect_navigation():
                    frame.click("input[name='btenvia.x']")

                frame.wait_for_selector("#vtabela")

                mensagem = frame.locator("div.box_mensagens").inner_text()

                resultado = "SUCESSO" if "Processo de integração foi iniciado" in mensagem else "ERRO"

                if resultado == "SUCESSO":
                    enviados += 1
                else:
                    erros += 1

                log(f"Resultado: {resultado} | Enviados:{enviados} | Erros:{erros}")

                salvar_xlsx(XLSX_PATH,{
                    "Codigo":codigo,
                    "Numero":numero,
                    "Cliente":cliente,
                    "Data Emissao":data_emissao,
                    "Resultado":resultado,
                    "Mensagem Retorno":mensagem.strip(),
                    "DataHoraExecucao":datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                })

                notas_processadas.add(chave)

                if not modo_primeira:

                    btn_ult = frame.locator("input[name='vpad-btult']")

                    if btn_ult.count() > 0:

                        btn_ult.click()

                        frame.wait_for_selector("input.radio_sel")

                break

            if not encontrou:

                if modo_primeira:

                    log("Processamento finalizado")
                    break

                btn_ant = frame.locator("input[name='vpad-btant']")

                if btn_ant.count() == 0:

                    log("Todas as páginas processadas")
                    break

                btn_ant.click()

                frame.wait_for_selector("input.radio_sel")

        browser.close()


if __name__ == "__main__":

    usuario = input("Usuário: ")
    senha = input("Senha: ")
    data_ini = input("Data início (dd/mm/yyyy): ")
    data_fim = input("Data fim (dd/mm/yyyy): ")

    executar_envio(usuario, senha, data_ini, data_fim)